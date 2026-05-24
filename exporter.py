# -*- coding: utf-8 -*-
"""Export du lieu tu SQLite ra file Excel theo cau truc Hop Dong."""
import os
import io
import shutil
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule

from db import get_conn, calc_summary

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
YELLOW_FILL = PatternFill("solid", start_color="FFF2CC")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
RED_FILL = PatternFill("solid", start_color="FFC7CE")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BOLD = Font(name="Arial", bold=True, size=11)
NORMAL = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def export_excel(output_path=None):
    conn = get_conn()
    wb = Workbook()

    # Sheet 1: Hop Dong
    ws1 = wb.active
    ws1.title = "1.DanhSachHopDong"
    h1 = ["Ma hop dong","Khach hang","Don vi thu huong","Phong","So hop dong","Gia tri HD","Tong giai ngan","% GN/HD",
          "Ngay ket thuc HD","Han bo sung HD cuoi","% Khau tru/dot",
          "Khau tru luy ke","HD da bo sung luy ke","Du can bo sung","% Hoan thanh","Trang thai","Ghi chu"]
    ws1.append(h1)
    _style_header(ws1, 1, len(h1))
    ws1.row_dimensions[1].height = 42
    for i,w in enumerate([16,22,22,12,15,16,16,11,14,16,13,16,16,16,13,16,22], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    rows = conn.execute("SELECT * FROM hop_dong ORDER BY created_at DESC").fetchall()
    today = date.today()
    for i, r in enumerate(rows, start=2):
        s = calc_summary(r["ma_hop_dong"])
        if not s: continue
        
        nkt = r["ngay_ket_thuc_hd"]
        nkt_date = datetime.strptime(nkt, "%Y-%m-%d").date() if isinstance(nkt, str) else nkt
        han_cuoi = nkt_date - timedelta(days=30) if nkt_date else None
        
        if s["tong_giai_ngan"] == 0: tt = "Chua giai ngan"
        elif s["hd_luy_ke"] >= s["tong_giai_ngan"]: tt = "Hoan tat"
        elif han_cuoi and today > han_cuoi: tt = "Qua han"
        elif han_cuoi and (han_cuoi - today).days <= 15: tt = "Sap den han"
        else: tt = "Dang theo doi"

        vals = [
            r["ma_hop_dong"], r["khach_hang"], r["don_vi_thu_huong"] or "", r["phong_phu_trach"] or "", r["so_hd"] or "",
            r["gia_tri_hd"] or 0, s["tong_giai_ngan"],
            (s["tong_giai_ngan"]/r["gia_tri_hd"]) if r["gia_tri_hd"] else 0,
            nkt_date, han_cuoi, r["pct_khau_tru"],
            s["khau_tru_luy_ke"], s["hd_luy_ke"], s["du_can_bo_sung"], s["pct_hoan_thanh"], tt, r["ghi_chu"] or ""
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws1.cell(row=i, column=c, value=v)
            cell.border = BORDER; cell.font = NORMAL
            cell.alignment = LEFT if c in (1,2,3,4,5,17) else CENTER
            
        for c in (6,7,12,13,14): ws1.cell(row=i,column=c).number_format = "#,##0;(#,##0);-"
        for c in (8,11,15): ws1.cell(row=i,column=c).number_format = "0.0%"
        for c in (9,10): ws1.cell(row=i,column=c).number_format = "dd/mm/yyyy"

    if rows:
        lr = 1 + len(rows)
        ws1.conditional_formatting.add(f"P2:P{lr}", FormulaRule(formula=['$P2="Qua han"'], fill=RED_FILL, font=Font(bold=True,color="9C0006")))
        ws1.conditional_formatting.add(f"P2:P{lr}", FormulaRule(formula=['$P2="Sap den han"'], fill=YELLOW_FILL, font=Font(bold=True,color="9C5700")))
        ws1.conditional_formatting.add(f"P2:P{lr}", FormulaRule(formula=['$P2="Hoan tat"'], fill=GREEN_FILL, font=Font(bold=True,color="006100")))
    ws1.freeze_panes = "B2"

    # Sheet 2: Giai Ngan
    ws2 = wb.create_sheet("2.CacDotGiaiNgan")
    h2 = ["Ma hop dong", "Ma giai ngan", "Ngay giai ngan", "So tien", "Ghi chu"]
    ws2.append(h2); _style_header(ws2, 1, len(h2))
    for i,w in enumerate([16,16,14,16,22], 1): ws2.column_dimensions[get_column_letter(i)].width = w
    
    gns = conn.execute("SELECT * FROM tam_ung ORDER BY ma_hop_dong, ngay_giai_ngan").fetchall()
    for i, r in enumerate(gns, start=2):
        vals = [r["ma_hop_dong"], r["ma_giai_ngan"], r["ngay_giai_ngan"], r["so_tien_tu"], r["ghi_chu"]]
        for c, v in enumerate(vals, start=1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.border = BORDER; cell.font = NORMAL
            cell.alignment = CENTER if c in (1,2,3) else LEFT
        ws2.cell(row=i,column=3).number_format = "dd/mm/yyyy"
        ws2.cell(row=i,column=4).number_format = "#,##0;(#,##0);-"
    ws2.freeze_panes = "B2"

    # Sheet 3: HSTT
    ws3 = wb.create_sheet("3.HoSoThanhToanKL")
    h3 = ["Ma hop dong","Dot #","Ngay HSTT","KL truoc VAT","% Khau tru","Khau tru dot","Ghi chu"]
    ws3.append(h3); _style_header(ws3,1,len(h3))
    for i,w in enumerate([16,8,14,18,11,18,22],1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    hstt = conn.execute("""
        SELECT h.*, t.pct_khau_tru as pct FROM hstt h
        LEFT JOIN hop_dong t ON h.ma_hop_dong = t.ma_hop_dong
        ORDER BY h.ma_hop_dong, h.dot_so
    """).fetchall()
    for i, r in enumerate(hstt, start=2):
        pct = r["pct"] or 0.1
        kt = r["kl_truoc_vat"] * pct
        vals = [r["ma_hop_dong"], r["dot_so"], r["ngay_hstt"], r["kl_truoc_vat"], pct, kt, r["ghi_chu"] or ""]
        for c, v in enumerate(vals, start=1):
            cell = ws3.cell(row=i, column=c, value=v)
            cell.border = BORDER; cell.font = NORMAL
            cell.alignment = LEFT if c == 7 else CENTER
        ws3.cell(row=i,column=3).number_format = "dd/mm/yyyy"
        for c in (4,6): ws3.cell(row=i,column=c).number_format = "#,##0;(#,##0);-"
        ws3.cell(row=i,column=5).number_format = "0.0%"
    ws3.freeze_panes = "B2"

    # Sheet 4: HD bo sung
    ws4 = wb.create_sheet("4.HoaDonBoSung")
    h4 = ["STT","Ma hop dong","Dot #","Ngay HD","So HD","MST ban","Ten ban","Tien truoc VAT","VAT","Tong","Status","Upload boi","Ngay upload","Ghi chu"]
    ws4.append(h4); _style_header(ws4,1,len(h4))
    for i,w in enumerate([6,16,8,12,14,14,28,16,14,16,11,14,14,20],1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    hd = conn.execute("SELECT * FROM hoa_don ORDER BY ma_hop_dong, dot_so, ngay_hd").fetchall()
    for i, r in enumerate(hd, start=2):
        vals = [i-1, r["ma_hop_dong"], r["dot_so"], r["ngay_hd"], r["so_hd"], r["mst_ban"] or "",
                r["ten_ban"] or "", r["tien_truoc_vat"], r["vat"], r["tong_cong"] or "",
                r["status"], r["uploaded_by"] or "", r["uploaded_at"], r["ghi_chu"] or ""]
        for c, v in enumerate(vals, start=1):
            cell = ws4.cell(row=i, column=c, value=v)
            cell.border = BORDER; cell.font = NORMAL
            cell.alignment = LEFT if c in (7,14) else CENTER
        ws4.cell(row=i,column=4).number_format = "dd/mm/yyyy"
        for c in (8,9,10): ws4.cell(row=i,column=c).number_format = "#,##0;(#,##0);-"
    ws4.freeze_panes = "B2"

    conn.close()

    if output_path is None:
        output_path = f"export_hopdong_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_path)
    return output_path


def export_to_drive_folder(drive_folder_path, filename_prefix="HopDong_TienDo"):
    today = date.today()
    fname = f"{filename_prefix}_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
    tmp_path = f"/tmp/{fname}" if os.name != 'nt' else fname
    export_excel(tmp_path)
    if not os.path.exists(drive_folder_path):
        os.makedirs(drive_folder_path, exist_ok=True)
    dest = os.path.join(drive_folder_path, fname)
    shutil.copy2(tmp_path, dest)
    latest = os.path.join(drive_folder_path, f"{filename_prefix}_LATEST.xlsx")
    shutil.copy2(tmp_path, latest)
    return dest, latest

# -*- coding: utf-8 -*-
"""
Web app theo doi tam ung & bo sung hoa don
Chay: streamlit run app.py --server.address 0.0.0.0 --server.port 8501
"""
import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta
import os
import io
import openpyxl
import copy
import db
import parsers
import exporter

st.set_page_config(page_title="Theo dõi Tạm ứng & HĐ", layout="wide", page_icon="📊")

# Init DB once
db.init_db()

# ============ AUTH ============
def login_screen():
    st.title("📊 Hệ thống theo dõi Tạm ứng & Bổ sung Hóa đơn")
    st.caption("Đăng nhập để tiếp tục")
    users = db.list_users()
    options = {f"{u['full_name']} ({u['role']} - {u['dept'] or ''})": u['username'] for u in users}
    choice = st.selectbox("Chọn user", list(options.keys()))
    
    username = options[choice]
    user_info = dict(db.get_user(username))
    
    password = st.text_input("Mật khẩu", type="password", placeholder="Nhập mật khẩu")
        
    if st.button("Đăng nhập", type="primary"):
        db_pass = user_info.get('password') or "123456"
        if password != db_pass:
            st.error("Mật khẩu không chính xác!")
        else:
            st.session_state["user"] = username
            st.session_state["user_info"] = user_info
            st.rerun()
    st.info("Demo: chon user de vao. Production: thay bang LDAP/SSO ngan hang.")


if "user" not in st.session_state:
    login_screen()
    st.stop()

user = st.session_state["user"]
uinfo = st.session_state["user_info"]
role = uinfo["role"]
is_qttd = role in ("qttd", "admin")
is_phong_kh = role == "phong_kh"

# ============ SIDEBAR ============
with st.sidebar:
    st.markdown(f"### 👤 {uinfo['full_name']}")
    st.caption(f"Role: **{role.upper()}** | Phòng: {uinfo['dept'] or '-'}")
    if st.button("Đăng xuất"):
        del st.session_state["user"]
        st.rerun()
    st.divider()

    if is_qttd:
        menu_items = [
            "📊 Dashboard",
        ]
        if user != "qttd01":
            menu_items.extend([
                "➕ Thêm Hợp đồng Tạm ứng",
                "📥 Import Hợp đồng (Excel)",
            ])
        menu_items.extend([
            "💼 Quản lý Hợp đồng",
            "✅ Duyệt HĐ chờ",
        ])
        if role == "admin":
            menu_items.append("📥 Up hóa đơn theo lô")
        menu_items.extend([
            "📦 Export & Drive",
            "🗂 Audit log",
        ])
        page = st.radio("Menu", menu_items)
    else:
        page = st.radio("Menu", [
            "📊 Dashboard (Của phòng)",
            "➕ Thêm Hợp đồng Tạm ứng",
            "💼 Danh sách Hợp đồng",
            "📥 Up hóa đơn theo lô",
        ])


# ============ HELPERS ============
def fmt_vnd(x):
    if x is None or x == "": return "-"
    try: return f"{int(float(x)):,}".replace(",", ".")
    except: return str(x)


def fmt_vnd_dot(x):
    """Format so tien voi dau cham ngan: 500.000.000"""
    if x is None or x == 0: return ""
    try: return f"{int(float(x)):,}".replace(",", ".")
    except: return str(x)


def parse_vnd_input(text):
    """Parse VND input: '500000000' hoac '500.000.000' -> float"""
    if not text: return None
    text = str(text).strip().replace(" ", "").replace(".", "").replace(",", "")
    try: return float(text)
    except: return None


def fmt_date(d):
    """Format date -> dd/mm/yyyy"""
    if isinstance(d, str):
        try: d = datetime.strptime(d, "%Y-%m-%d").date()
        except: return d
    if isinstance(d, date): return d.strftime("%d/%m/%Y")
    return str(d) if d else "-"


def status_badge(s):
    colors = {"Hoàn tất": "green", "Quá hạn": "red", "Sắp đến hạn": "orange",
              "Đang theo dõi": "blue", "Du HD": "green", "Tre han": "red",
              "Dang cho HD": "orange", "approved": "green", "pending": "orange", "rejected": "red"}
    return f":{colors.get(s, 'gray')}[**{s}**]"

def status_badge_html(s):
    colors = {"Hoàn tất": "#198754", "Quá hạn": "#dc3545", "Sắp đến hạn": "#fd7e14",
              "Đang theo dõi": "#0d6efd", "approved": "#198754", "pending": "#fd7e14", "rejected": "#dc3545"}
    c = colors.get(s, 'gray')
    return f"<span style='color:{c}; font-weight:bold;'>{s}</span>"


def shorten_company_name(name):
    if not name: return ""
    s = str(name).upper()
    prefixes = [
        "CÔNG TY", "CTCP", "CỔ PHẦN", "TRÁCH NHIỆM HỮU HẠN", "TNHH", "MTV", 
        "THƯƠNG MẠI", "TM", "DỊCH VỤ", "DV", "ĐẦU TƯ", "XÂY DỰNG", 
        "PHÁT TRIỂN", "TẬP ĐOÀN", "VÀ", "&", "VẬT LIỆU", "SẢN XUẤT", 
        "KINH DOANH", "XUẤT NHẬP KHẨU", "CỔ PHẦN"
    ]
    for p in prefixes:
        s = s.replace(p, "")
    s = " ".join(s.split()) # normalize spaces
    if not s: return str(name)[:15] # fallback
    return (s[:15] + '..') if len(s) > 15 else s


def compute_hd_status(r):
    today = date.today()
    summary = db.calc_summary(r.get("ma_hop_dong"))
    if not summary: return summary, "?"
    nkt = r.get("ngay_ket_thuc_hd")
    if isinstance(nkt, str):
        nkt = datetime.strptime(nkt, "%Y-%m-%d").date()
    # Hạn bổ sung HĐ = Ngày kết thúc HĐ + 30 ngày (theo CV 4483)
    han_bo_sung = nkt + timedelta(days=30)
    
    if summary["tong_giai_ngan"] == 0:
        tt = "Chưa giải ngân"
    elif summary["hd_luy_ke"] >= summary["tong_giai_ngan"]: 
        tt = "Hoàn tất"
    elif today > han_bo_sung: tt = "Quá hạn"
    elif (han_bo_sung - today).days <= 15: tt = "Sắp đến hạn"
    else: tt = "Đang theo dõi"
    
    summary["han_cuoi"] = han_bo_sung
    # Tạm ứng còn lại = Tổng GN - Tổng khấu trừ HSTT
    summary["tam_ung_con_lai"] = max(summary["tong_giai_ngan"] - summary["khau_tru_luy_ke"], 0)
    return summary, tt


# ============ PAGE: DASHBOARD ============
def page_dashboard():
    st.title("📊 Dashboard")
    phong = uinfo["dept"] if is_phong_kh else None
    hd_list = db.list_hop_dong(phong=phong)

    # --- Filters ---
    kh_list = sorted(list(set([r.get("khach_hang") for r in hd_list if r.get("khach_hang")])))
    dvth_list = sorted(list(set([r.get("don_vi_thu_huong") for r in hd_list if r.get("don_vi_thu_huong")])))
    phong_list = sorted(list(set([r.get("phong_phu_trach") for r in hd_list if r.get("phong_phu_trach")])))
    
    with st.expander("🔍 Lọc dữ liệu", expanded=False):
        c1, c2 = st.columns(2)
        kh_filters = c1.multiselect("Khách hàng", kh_list)
        if is_qttd:
            phong_filters = c2.multiselect("Phòng quản lý", phong_list)
            dvth_filters = []
        else:
            dvth_filters = c2.multiselect("Đơn vị thụ hưởng", dvth_list)
            phong_filters = []
        
    if kh_filters:
        hd_list = [r for r in hd_list if r.get("khach_hang") in kh_filters]
    if dvth_filters:
        hd_list = [r for r in hd_list if r.get("don_vi_thu_huong") in dvth_filters]
    if phong_filters:
        hd_list = [r for r in hd_list if r.get("phong_phu_trach") in phong_filters]

    total_gn = 0
    total_hd = 0
    qua_han = sap_han = hoan_tat = 0
    rows_display = []
    
    for r in hd_list:
        s, tt = compute_hd_status(r)
        total_gn += s["tong_giai_ngan"]
        total_hd += s["hd_luy_ke"]
        if tt == "Quá hạn": qua_han += 1
        elif tt == "Sắp đến hạn": sap_han += 1
        elif tt == "Hoàn tất": hoan_tat += 1
        loai = r.get("loai_tu")
        rows_display.append({
            "Ngày giải ngân": fmt_date(s.get("ngay_giai_ngan_dau")),
            "Mã giải ngân": r.get("ma_hop_dong"), "Khách hàng": r.get("khach_hang"),
            "Số hợp đồng": r.get("so_hd") or "-",
            "ĐV thụ hưởng": r.get("don_vi_thu_huong") or "",
            "Loại": "Từng đợt" if loai == "khau_tru_dot" else "1 lần",
            "Phòng": r.get("phong_phu_trach") or "",
            "Tổng giải ngân": s.get("tong_giai_ngan", 0), "HĐ bổ sung": s.get("hd_luy_ke", 0),
            "Dư cần bổ sung": s.get("du_can_bo_sung", 0),
            "Hạn bổ sung": fmt_date(s.get("han_cuoi")),
            "% Hoàn thành": s.get("pct_hoan_thanh", 0), "Trạng thái": tt,
        })

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Tổng số Hợp đồng", len(hd_list))
    c2.metric("Đã giải ngân (tr)", f"{total_gn/1e6:,.0f}")
    c3.metric("Đã thu HĐ (tr)", f"{total_hd/1e6:,.0f}")
    c4.metric("Tỷ lệ thu HĐ", f"{(total_hd/total_gn*100 if total_gn else 0):.1f}%")

    c1, c2, c3 = st.columns(3)
    c1.metric("🔴 Quá hạn", qua_han)
    c2.metric("🟡 Sắp đến hạn", sap_han)
    c3.metric("🟢 Hoàn tất", hoan_tat)

    # --- Cảnh báo quá hạn / sắp hạn ---
    alerts_qh = [r for r in rows_display if r["Trạng thái"] == "Quá hạn"]
    alerts_sh = [r for r in rows_display if r["Trạng thái"] == "Sắp đến hạn"]
    if alerts_qh:
        with st.expander(f"🚨 Cảnh báo: Có {len(alerts_qh)} Hợp đồng QUÁ HẠN bổ sung hóa đơn", expanded=False):
            for a in alerts_qh:
                st.error(f"**{a['Khách hàng']}** — HĐ: {a['Số hợp đồng']} — Hạn BS: {a['Hạn bổ sung']} — Dư cần BS: {fmt_vnd(a['Dư cần bổ sung'])}")
    if alerts_sh:
        with st.expander(f"⚠️ Cảnh báo: Có {len(alerts_sh)} Hợp đồng SẮP ĐẾN HẠN bổ sung hóa đơn", expanded=False):
            for a in alerts_sh:
                st.warning(f"**{a['Khách hàng']}** — HĐ: {a['Số hợp đồng']} — Hạn BS: {a['Hạn bổ sung']} — Dư cần BS: {fmt_vnd(a['Dư cần bổ sung'])}")

    if rows_display:
        df = pd.DataFrame(rows_display)
        df["Tổng giải ngân"] = df["Tổng giải ngân"].apply(fmt_vnd)
        df["HĐ bổ sung"] = df["HĐ bổ sung"].apply(fmt_vnd)
        df["Dư cần bổ sung"] = df["Dư cần bổ sung"].apply(fmt_vnd)
        df["% Hoàn thành"] = df["% Hoàn thành"].apply(lambda x: f"{x*100:.1f}%")
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("Chưa có hợp đồng nào.")


# ============ PAGE: QUAN LY HOP DONG ============
def page_hop_dong():
    st.title("💼 Quản lý Hợp đồng & Bổ sung Hóa đơn")

    st.subheader("🔍 Tìm kiếm")
    fc1, fc2, fc3 = st.columns([3, 1, 1])
    search_text = fc1.text_input("Tìm kiếm chung", placeholder="Nhập tên KH, Đơn vị thụ hưởng, mã HĐ, số HĐ...")
    filter_phong = fc2.selectbox("Phòng", ["Tất cả", "PGD Hàng Xanh", "PGD Đinh Tiên Hoàng", "PGD Đakao", "PGD Nguyễn Oanh", "PGD Tân Thới Hiệp", "KHDN1", "KHDN2", "KH FDI"])
    filter_loai = fc3.selectbox("Loại", ["Tất cả", "Khấu trừ từng đợt", "Thanh toán 1 lần"])

    phong_filter = None if is_qttd else uinfo["dept"]
    all_rows = db.list_hop_dong(phong=phong_filter)
    filtered = []
    for r in all_rows:
        loai = r.get("loai_tu")
        if filter_phong != "Tất cả" and r.get("phong_phu_trach") != filter_phong: continue
        if filter_loai == "Khấu trừ từng đợt" and loai != "khau_tru_dot": continue
        if filter_loai == "Thanh toán 1 lần" and loai != "mot_lan": continue
        if search_text:
            q = search_text.strip().upper()
            searchable = f"{r.get('ma_hop_dong')} {r.get('khach_hang')} {r.get('don_vi_thu_huong')} {r.get('so_hd')} {r.get('ghi_chu')}".upper()
            if q not in searchable: continue
            
        # Tinh toan summary de su dung trong dropdown hien thi
        s, tt = compute_hd_status(r)
        
        # Format string cho dropdown: BCONS-số HĐ-số tiền-đơn vị thụ hưởng
        kh_short = shorten_company_name(r.get('khach_hang', ''))
        dv_short = shorten_company_name(r.get('don_vi_thu_huong', ''))
        so_hd = r.get('so_hd') or "Chưa rõ"
        tien = f"{fmt_vnd_dot(s['tong_giai_ngan'])}đ" if s['tong_giai_ngan'] else "0đ"
        
        display_name = f"{kh_short} - HĐ: {so_hd} - GN: {tien}"
        if dv_short: display_name += f" - ĐVTH: {dv_short}"
        
        filtered.append({
            "data": r,
            "summary": s,
            "tt": tt,
            "display": display_name
        })

    if not filtered:
        st.info("Không tìm thấy hợp đồng nào.")
        return

    selected_idx = st.selectbox(
        "📌 Chọn Hợp đồng để xem chi tiết & bổ sung hóa đơn:",
        range(len(filtered)),
        format_func=lambda i: filtered[i]["display"]
    )
    
    selected_item = filtered[selected_idx]
    selected = selected_item["data"]
    s = selected_item["summary"]
    tt = selected_item["tt"]
    loai = selected.get("loai_tu")

    if loai == "khau_tru_dot":
        tab_names = ["📄 Chi tiết HĐ", "💸 Các đợt Giải ngân", "📋 HSTT", "📤 Bổ sung HĐ", "🧾 DS Hóa đơn"]
    else:
        tab_names = ["📄 Chi tiết HĐ", "💸 Các đợt Giải ngân", "📤 Bổ sung HĐ", "🧾 DS Hóa đơn"]
        
    tabs = st.tabs(tab_names)

    # --- TAB: Chi tiet HD ---
    with tabs[0]:
        c_title, c_btn = st.columns([4, 1])
        c_title.markdown("#### Thông tin Hợp đồng")
        if uinfo["role"] == "admin":
            if c_btn.button("🗑️ Xóa Hợp đồng này", type="primary"):
                db.delete_hop_dong(selected.get('ma_hop_dong'), user)
                st.rerun()
        c1, c2, c3, c4 = st.columns(4)
        khe_uoc = selected.get('khe_uoc_vay') or '-'
        c1.markdown(f"**Mã giải ngân:** `{selected.get('ma_hop_dong')}`<br>**Khế ước vay:** {khe_uoc}", unsafe_allow_html=True)
        c2.markdown(f"**Khách hàng:** {selected.get('khach_hang')} " + (f"(CIF: `{selected.get('cif')}`)" if selected.get('cif') else ""))
        c3.markdown(f"**Trạng thái:** {status_badge(tt)}")
        c4.markdown(f"**Loại:** {'Khấu trừ từng đợt' if loai == 'khau_tru_dot' else 'Thanh toán 1 lần'}")

        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(f"**Số hợp đồng:** {selected.get('so_hd') or '-'}")
        ngay_hd_val = selected.get('ngay_hop_dong')
        c2.markdown(f"**Ngày hợp đồng:** {fmt_date(ngay_hd_val) if ngay_hd_val else '-'}")
        c3.markdown(f"**ĐV thụ hưởng:** {selected.get('don_vi_thu_huong') or '-'}")
        c4.markdown(f"**Ngày kết thúc hợp đồng:** {fmt_date(selected.get('ngay_ket_thuc_hd'))}")
        
        c_p = st.container()
        c_p.markdown(f"**Phòng phụ trách:** {selected.get('phong_phu_trach') or '-'}")
        
        if selected.get("ghi_chu"):
            st.markdown(f"**Ghi chú:** {selected.get('ghi_chu')}")

        st.divider()
        st.markdown("#### Số liệu tài chính")
        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(f"<div style='background-color:#f0f2f6;padding:10px;border-radius:5px;'><b>Giá trị hợp đồng</b><br><span style='font-size:1.4rem;color:#1F4E78'>{fmt_vnd_dot(selected.get('gia_tri_hd')) + ' ₫' if selected.get('gia_tri_hd') else '-'}</span></div>", unsafe_allow_html=True)
        c2.markdown(f"<div style='background-color:#e8f4f8;padding:10px;border-radius:5px;'><b>Tổng Giải Ngân</b><br><span style='font-size:1.4rem;color:#0d6efd'>{fmt_vnd_dot(s['tong_giai_ngan'])} ₫</span></div>", unsafe_allow_html=True)
        c3.markdown(f"<div style='background-color:#e6f4ea;padding:10px;border-radius:5px;'><b>Đã thu HĐ</b><br><span style='font-size:1.4rem;color:#198754'>{fmt_vnd_dot(s['hd_luy_ke'])} ₫</span></div>", unsafe_allow_html=True)
        c4.markdown(f"<div style='background-color:#fce8e6;padding:10px;border-radius:5px;'><b>Còn phải thu</b><br><span style='font-size:1.4rem;color:#dc3545'>{fmt_vnd_dot(s['du_can_bo_sung'])} ₫</span></div>", unsafe_allow_html=True)
        
        c1, c2 = st.columns(2)
        han_bs = fmt_date(s.get("han_cuoi"))
        c1.markdown(f"<div style='background-color:#e3f2fd;padding:10px;border-radius:5px;'><b>Hạn bổ sung HĐ</b><br><span style='font-size:1.4rem;color:#1565c0'>{han_bs}</span></div>", unsafe_allow_html=True)
        c2.markdown(f"<div style='background-color:#f3e5f5;padding:10px;border-radius:5px;'><b>Trạng thái</b><br><span style='font-size:1.4rem;'>{status_badge_html(tt)}</span></div>", unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        st.progress(min(s["pct_hoan_thanh"], 1.0), text=f"Hoàn thành: {s['pct_hoan_thanh']*100:.1f}%")



    # --- TAB: Giai ngan ---
    with tabs[1]:
        gns = db.list_tam_ung(selected.get("ma_hop_dong"))
        if gns:
            df_gn = pd.DataFrame([{
                "Mã giải ngân": g["ma_giai_ngan"],
                "Ngày giải ngân": fmt_date(g["ngay_giai_ngan"]),
                "Số tiền": fmt_vnd_dot(g["so_tien_tu"]),
                "Ghi chú": g["ghi_chu"] or "",
                "Người tạo": g["created_by"]
            } for g in gns])
            
            edited_df = st.data_editor(
                df_gn, 
                use_container_width=True, 
                hide_index=True,
                disabled=["Mã giải ngân", "Ngày giải ngân", "Số tiền", "Người tạo"],
                key=f"editor_gn_{selected.get('ma_hop_dong')}"
            )
            
            changed = False
            for i in range(len(df_gn)):
                if df_gn.iloc[i]["Ghi chú"] != edited_df.iloc[i]["Ghi chú"]:
                    db.update_tam_ung_ghi_chu(df_gn.iloc[i]["Mã giải ngân"], edited_df.iloc[i]["Ghi chú"])
                    changed = True
            if changed:
                st.rerun()
        else:
            st.info("Chưa có đợt giải ngân nào.")
            
        with st.expander("➕ Thêm đợt giải ngân mới cho HĐ này"):
            with st.form(f"add_gn_{selected.get('ma_hop_dong')}"):
                c1, c2 = st.columns(2)
                stien_txt = c1.text_input("Số tiền giải ngân (VND) *", placeholder="VD: 50.000.000")
                ngay_gn = c2.date_input("Ngày giải ngân", value=date.today(), format="DD/MM/YYYY")
                ghi_chu = st.text_input("Ghi chú")
                stien = parse_vnd_input(stien_txt)
                if stien:
                    st.caption(f"= {fmt_vnd_dot(stien)} VND")
                    
                if st.form_submit_button("Lưu đợt giải ngân", type="primary"):
                    if not stien:
                        st.error("Vui lòng nhập số tiền hợp lệ!")
                    else:
                        ma_gn = db.generate_ma_giai_ngan(selected.get("khach_hang"), year=ngay_gn.year)
                        db.add_tam_ung({
                            "ma_giai_ngan": ma_gn, "ma_hop_dong": selected.get("ma_hop_dong"),
                            "so_tien_tu": stien, "ngay_giai_ngan": ngay_gn, "ghi_chu": ghi_chu
                        }, user)
                        st.success(f"Đã thêm đợt giải ngân {ma_gn}")
                        st.rerun()

    tab_offset = 1
    # --- TAB: HSTT ---
    if loai == "khau_tru_dot":
        with tabs[2]:
            hstt_rows = db.list_hstt(selected.get("ma_hop_dong"))
            if hstt_rows:
                data = []
                kt_luy_ke = 0
                for r in hstt_rows:
                    loai_kt = selected.get("loai_gia_tri_kt") if selected.get("loai_gia_tri_kt") else "Trước VAT"
                    # r is a dict, so r.get works
                    gia_tri_tinh_kt = r["tong_cong"] if loai_kt == "Sau VAT" and r["tong_cong"] is not None else r["kl_truoc_vat"]
                    kt = gia_tri_tinh_kt * (selected.get("pct_khau_tru") / 100.0)
                    data.append({
                        "Đợt": r["dot_so"],
                        "Ngày HSTT": fmt_date(r["ngay_hstt"]),
                        "KL trước VAT": fmt_vnd(r["kl_truoc_vat"]),
                        "VAT": fmt_vnd(r["vat"] if r["vat"] is not None else 0),
                        "Tổng cộng": fmt_vnd(r["tong_cong"] if r["tong_cong"] is not None else r["kl_truoc_vat"]),
                        "Khấu trừ TU": fmt_vnd(kt),
                        "Ghi chú": r["ghi_chu"] or "",
                    })
                st.dataframe(pd.DataFrame(data), use_container_width=True, hide_index=True)
                
                if role.lower() == "admin" and hstt_rows:
                    with st.expander("✏️ Chỉnh sửa / Xóa HSTT"):
                        edit_options = [f"Đợt {r['dot_so']} - {fmt_date(r['ngay_hstt'])}" for r in hstt_rows]
                        edit_choice = st.selectbox("Chọn Đợt HSTT cần sửa:", edit_options)
                        idx = edit_options.index(edit_choice)
                        edit_r = hstt_rows[idx]
                        
                        with st.form(f"edit_hstt_form_{selected.get('ma_hop_dong')}"):
                            c1, c2 = st.columns(2)
                            ed_dot = c1.number_input("Sửa Đợt #", min_value=1, value=edit_r.get("dot_so"))
                            
                            ed_ngay_str = edit_r.get("ngay_hstt")
                            try:
                                if isinstance(ed_ngay_str, str): ed_ngay_dt = datetime.strptime(ed_ngay_str, "%Y-%m-%d").date()
                                else: ed_ngay_dt = ed_ngay_str
                            except: ed_ngay_dt = date.today()
                            
                            ed_ngay = c2.date_input("Sửa Ngày HSTT", value=ed_ngay_dt, format="DD/MM/YYYY")
                            
                            c1, c2 = st.columns(2)
                            ed_kl_text = c1.text_input("KL trước VAT (VND) *", value=f"{int(edit_r.get('kl_truoc_vat')):,}".replace(",", "."))
                            ed_vat_val = edit_r.get('vat') if edit_r.get('vat') is not None else 0
                            ed_vat_text = c2.text_input("VAT (VND)", value=f"{int(ed_vat_val):,}".replace(",", "."))
                            ed_ghi_chu = st.text_input("Ghi chú", value=edit_r.get("ghi_chu") or "")
                            
                            col1, col2 = st.columns(2)
                            if col1.form_submit_button("💾 Lưu thay đổi", type="primary", use_container_width=True):
                                ed_kl = parse_vnd_input(ed_kl_text)
                                ed_vat = parse_vnd_input(ed_vat_text) or 0
                                if not ed_kl or ed_kl <= 0:
                                    st.error("Nhập KL hợp lệ!")
                                else:
                                    db.update_hstt(edit_r.get("id"), {
                                        "dot_so": int(ed_dot), "ngay_hstt": ed_ngay,
                                        "kl_truoc_vat": ed_kl, "vat": ed_vat,
                                        "tong_cong": ed_kl + ed_vat, "ghi_chu": ed_ghi_chu
                                    }, user)
                                    st.success("Đã cập nhật!")
                                    st.rerun()
                            if col2.form_submit_button("❌ Xóa đợt này", use_container_width=True):
                                db.delete_hstt(edit_r.get("id"), user)
                                st.success("Đã xóa!")
                                st.rerun()
            else:
                st.info("Chưa có HSTT nào.")

            with st.expander("➕ Thêm đợt HSTT mới"):
                with st.form(f"hstt_{selected.get('ma_hop_dong')}"):
                    next_dot = max([r["dot_so"] for r in hstt_rows], default=0) + 1
                    c1, c2 = st.columns(2)
                    dot = c1.number_input("Đợt #", min_value=1, value=next_dot)
                    ngay = c2.date_input("Ngày HSTT", value=date.today(), format="DD/MM/YYYY")
                    
                    c1, c2 = st.columns(2)
                    kl_text = c1.text_input("KL trước VAT (VND) *", placeholder="VD: 1.000.000.000")
                    vat_text = c2.text_input("VAT (VND)", placeholder="VD: 100.000.000")
                    ghi_chu = st.text_input("Ghi chú")
                    
                    kl = parse_vnd_input(kl_text)
                    vat_amt = parse_vnd_input(vat_text) or 0
                    
                    if kl:
                        tong_cong = kl + vat_amt
                        loai_kt = selected.get("loai_gia_tri_kt") if selected.get("loai_gia_tri_kt") else "Trước VAT"
                        gia_tri_tinh_kt = tong_cong if loai_kt == "Sau VAT" else kl
                        kt = gia_tri_tinh_kt * (selected.get('pct_khau_tru') / 100.0)
                        st.caption(f"= KL: {fmt_vnd_dot(kl)} + VAT: {fmt_vnd_dot(vat_amt)} = Tổng: {fmt_vnd_dot(tong_cong)}")
                        st.caption(f"→ Khấu trừ ({loai_kt}): {fmt_vnd_dot(kt)} VND")
                        
                    if st.form_submit_button("Thêm HSTT", type="primary"):
                        if not kl or kl <= 0:
                            st.error("Nhập KL hợp lệ!")
                        else:
                            tong_cong = kl + vat_amt
                            db.add_hstt({"ma_hop_dong": selected.get("ma_hop_dong"), "dot_so": int(dot),
                                         "ngay_hstt": ngay, "kl_truoc_vat": kl, "vat": vat_amt, 
                                         "tong_cong": tong_cong, "ghi_chu": ghi_chu}, user)
                            st.rerun()
        tab_offset = 2

    # --- TAB: Bo sung HD ---
    with tabs[1 + tab_offset]:
        st.subheader("Upload hóa đơn bổ sung")
        files = st.file_uploader("Chọn file XML/PDF", accept_multiple_files=True, type=["xml", "pdf"], key=f"up_{selected.get('ma_hop_dong')}")
        
        if loai == "khau_tru_dot":
            hstt_list = db.list_hstt(selected.get("ma_hop_dong"))
            max_dot = max([r["dot_so"] for r in hstt_list], default=1)
            dot_sel = st.number_input("Gán vào đợt #", min_value=1, value=max_dot, step=1, key=f"dot_{selected.get('ma_hop_dong')}")
        else:
            dot_sel = 1

        if files and st.button("📥 Trích xuất file"):
            if not os.path.exists("uploads"):
                os.makedirs("uploads")
            ok = 0
            for f in files:
                file_path = os.path.join("uploads", f.name)
                with open(file_path, "wb") as out_f:
                    out_f.write(f.getvalue())
                
                f.seek(0)
                d = parsers.parse_file(f)
                d["ma_hop_dong"] = selected.get("ma_hop_dong")
                d["dot_so"] = int(dot_sel)
                d["file_src"] = file_path
                db.add_staging(d, user)
                ok += 1
            st.success(f"Đã trích xuất {ok} file. Vui lòng Xác nhận bên dưới!")
            st.rerun()

        stagings = [s for s in db.list_staging(user=user if is_phong_kh else None) if s["ma_hop_dong"] == selected.get("ma_hop_dong")]
        if stagings:
            st.markdown("##### 📝 Hóa đơn chờ xác nhận")
            for s in stagings:
                with st.expander(f"HĐ: {s['so_hd'] or 'Chưa rõ'} - {s['file_src'] or ''}", expanded=True):
                    with st.form(f"cf_{s['id']}"):
                        c1, c2 = st.columns(2)
                        so_hd = c1.text_input("Số hóa đơn *", value=s['so_hd'] or "")
                        def_date = date.today()
                        if s['ngay_hd']:
                            try:
                                def_date = datetime.strptime(s["ngay_hd"], "%Y-%m-%d").date() if isinstance(s["ngay_hd"], str) else s["ngay_hd"]
                            except: pass
                        ngay_hd = c2.date_input("Ngày hóa đơn", value=def_date, format="DD/MM/YYYY")
                        
                        c1, c2 = st.columns(2)
                        ten_ban = c1.text_input("Tên người bán", value=s['ten_ban'] or "")
                        mst = c2.text_input("MST người bán", value=s['mst_ban'] or "")
                        
                        tien_text = c1.text_input("Tiền trước VAT *", value=f"{int(s['tien_truoc_vat'])}" if s['tien_truoc_vat'] is not None else "")
                        vat_text = c2.text_input("Tiền VAT", value=f"{int(s['vat'])}" if s['vat'] is not None else "")
                        
                        tien = parse_vnd_input(tien_text)
                        vat = parse_vnd_input(vat_text) or 0
                        
                        dvth = selected.get('don_vi_thu_huong') or ""
                        norm_dvth = db.normalize_company_name(dvth) if dvth else ""
                        norm_tb = db.normalize_company_name(ten_ban) if ten_ban else ""
                        
                        mismatch = False
                        if norm_dvth and norm_tb and norm_dvth != norm_tb:
                            st.warning(f"⚠️ Cảnh báo: Tên người bán ({ten_ban}) KHÔNG KHỚP với Đơn vị thụ hưởng ({dvth})!")
                            mismatch = True
                            confirm_mismatch = st.checkbox("Xác nhận Hóa đơn hợp lệ dù tên không khớp", key=f"chk_{s['id']}")
                        else:
                            confirm_mismatch = True
                        
                        c1, c2, c3 = st.columns([2, 1, 1])
                        if c2.form_submit_button("✓ Xác nhận lưu", type="primary"):
                            if mismatch and not confirm_mismatch:
                                st.error("Vui lòng tick chọn Xác nhận hợp lệ ở bên trên!")
                            elif not so_hd or not tien:
                                st.error("Nhập số HĐ và tiền!")
                            else:
                                db.add_hoa_don({
                                    "ma_hop_dong": selected.get("ma_hop_dong"), "dot_so": s["dot_so"],
                                    "so_hd": so_hd, "ngay_hd": ngay_hd, "mst_ban": mst, "ten_ban": ten_ban,
                                    "tien_truoc_vat": tien, "vat": vat, "tong_cong": tien + vat,
                                    "ma_tra_cuu": s['ma_tra_cuu'], "file_src": s['file_src'], "ghi_chu": f"Upload boi {user}"
                                }, user, status="approved" if is_qttd else "pending")
                                db.delete_staging(s["id"])
                                st.rerun()
                        if c3.form_submit_button("🗑️ Xóa file"):
                            db.delete_staging(s["id"])
                            st.rerun()

        with st.expander("✍️ Nhập thủ công hóa đơn"):
            with st.form(f"mn_{selected.get('ma_hop_dong')}"):
                c1, c2 = st.columns(2)
                m_hd = c1.text_input("Số hóa đơn *")
                m_dt = c2.date_input("Ngày hóa đơn", value=date.today(), format="DD/MM/YYYY")
                m_tien = c1.text_input("Tiền trước VAT *")
                m_vat = c2.text_input("Tiền VAT")
                
                if st.form_submit_button("Lưu hóa đơn", type="primary"):
                    tien = parse_vnd_input(m_tien)
                    vat = parse_vnd_input(m_vat) or 0
                    if not m_hd or not tien:
                        st.error("Lỗi dữ liệu")
                    else:
                        db.add_hoa_don({
                            "ma_hop_dong": selected.get("ma_hop_dong"), "dot_so": int(dot_sel),
                            "so_hd": m_hd, "ngay_hd": m_dt, "tien_truoc_vat": tien, "vat": vat,
                            "tong_cong": tien + vat, "file_src": "manual"
                        }, user, status="approved" if is_qttd else "pending")
                        st.rerun()

    # --- TAB: DS Hoa don ---
    with tabs[2 + tab_offset]:
        hd_list = db.list_hoa_don(ma_hop_dong=selected.get("ma_hop_dong"))
        if hd_list:
            pct = 1.0 if selected.get("loai_tu") == "mot_lan" else (selected.get("pct_khau_tru") or 0) / 100.0
            loai_kt = "Trước VAT" if selected.get("loai_tu") == "mot_lan" else (selected.get("loai_gia_tri_kt") or "Trước VAT")
            
            for h in hd_list:
                gia_tri_tinh = h['tong_cong'] if loai_kt == "Sau VAT" else h['tien_truoc_vat']
                can_tru = (gia_tri_tinh or 0) * pct
                
                with st.container(border=True):
                    c1, c2 = st.columns([3, 1])
                    with c1:
                        uploaded = h['uploaded_at'] if 'uploaded_at' in h.keys() else None
                        ngay_bs = fmt_date(str(uploaded).split(' ')[0]) if uploaded else '-'
                        st.write(f"**HD {h['so_hd']}** | Ngày HĐ: {fmt_date(h['ngay_hd'])} | Ngày BS: {ngay_bs} | Đợt: {h['dot_so']} | TT: {status_badge(h['status'])}")
                        st.write(f"**Người bán:** {h['ten_ban'] or '-'} | MST: `{h['mst_ban'] or '-'}`")
                        st.write(f"**Số tiền:** {fmt_vnd(h['tien_truoc_vat'])} | **VAT:** {fmt_vnd(h['vat'])} | **Tổng:** {fmt_vnd(h['tong_cong'])}")
                        st.markdown(f"Tương đương Giá trị thu hồi tạm ứng: **<span style='color:green;'>{fmt_vnd(can_tru)} VND</span>** *(Tính theo {pct*100:g}% của {loai_kt})*", unsafe_allow_html=True)
                        if h['file_src'] and os.path.exists(h['file_src']):
                            with open(h['file_src'], "rb") as f:
                                st.download_button("📥 Tải file đính kèm", f, file_name=os.path.basename(h['file_src']), key=f"dl_hd_{h['id']}")
                    with c2:
                        if st.button("🗑️ Xóa", key=f"del_{h['id']}", use_container_width=True):
                            db.delete_hoa_don(h['id'], user)
                            st.rerun()
        else:
            st.info("Chưa có hóa đơn nào.")


# ============ PAGE: THEM HOP DONG MOI ============
def page_them_hop_dong():
    st.title("➕ Thêm Hợp đồng Tạm ứng mới")
    loai_tu = st.radio("Loại thanh toán *", ["Khấu trừ từng đợt", "Thanh toán 1 lần"], horizontal=True)
    is_mot_lan = loai_tu == "Thanh toán 1 lần"
    companies = db.list_distinct_companies()

    with st.container():
        c1, c2 = st.columns(2)
        kh_select = c1.selectbox("Chọn KH đã có", ["--- Nhập mới ---"] + companies) if companies else "--- Nhập mới ---"
        kh_new = c2.text_input("Hoặc nhập tên mới")
        
        cif_text = st.text_input("Mã CIF Khách hàng")
        
        c1, c2 = st.columns(2)
        dv_select = c1.selectbox("Đơn vị thụ hưởng", ["--- Nhập mới ---"] + companies) if companies else "--- Nhập mới ---"
        dv_new = c2.text_input("Hoặc nhập ĐVTH mới")
        
        c1, c2, c3 = st.columns(3)
        so_hd = c1.text_input("Số hợp đồng *")
        ngay_hd = c2.date_input("Ngày hợp đồng", value=date.today(), format="DD/MM/YYYY")
        phong = uinfo["dept"] if is_phong_kh else c3.selectbox("Phòng phụ trách", ["PGD Hàng Xanh", "PGD Đinh Tiên Hoàng", "PGD Đakao", "PGD Nguyễn Oanh", "PGD Tân Thới Hiệp", "KHDN1", "KHDN2", "KH FDI"])
        
        def format_currency(key):
            val = st.session_state.get(key, "")
            if val:
                num = parse_vnd_input(val)
                if num and num > 0:
                    if num.is_integer():
                        st.session_state[key] = f"{int(num):,}".replace(",", ".")
                    else:
                        st.session_state[key] = f"{num:,}".replace(",", ".")
                else:
                    st.session_state[key] = ""

        c1, c2 = st.columns(2)
        if "gtri_input" not in st.session_state:
            st.session_state["gtri_input"] = ""
        if "tu_input" not in st.session_state:
            st.session_state["tu_input"] = ""

        gtri_text = c1.text_input("Giá trị HĐ (VND)", key="gtri_input", on_change=format_currency, args=("gtri_input",))
        tu_text = c2.text_input("Số tiền tạm ứng Đợt 1 *", key="tu_input", on_change=format_currency, args=("tu_input",))
        
        c1, c2 = st.columns(2)
        ngay_gn = c1.date_input("Ngày giải ngân đợt 1", value=date.today(), format="DD/MM/YYYY")
        ngay_kt = c2.date_input("Ngày kết thúc HĐ", value=date.today() + timedelta(days=180), format="DD/MM/YYYY")
        
        c1, c2 = st.columns(2)
        pct = 0.0 if is_mot_lan else c1.number_input("% Khấu trừ/đợt", value=10.0, step=1.0)
        loai_kt = "" if is_mot_lan else c2.radio("Tính khấu trừ theo", ["Trước VAT", "Sau VAT"], horizontal=True)
        
        khe_uoc_vay = st.text_input("Số Khế ước vay")
        ghi_chu = st.text_area("Ghi chú")
        
        gtri = parse_vnd_input(gtri_text)
        tu_amt = parse_vnd_input(tu_text)

        if st.button("➕ Lưu Hợp đồng & Giải ngân đợt 1", type="primary"):
            kh = kh_new.strip() or (kh_select if kh_select != "--- Nhập mới ---" else "")
            dvth = dv_new.strip() or (dv_select if dv_select != "--- Nhập mới ---" else "")
            
            if not kh or not tu_amt:
                st.error("Nhập đầy đủ Khách hàng và Số tiền!")
            else:
                ma_hd = db.generate_ma_hop_dong(kh, year=ngay_gn.year)
                # Kiem tra neu hop dong da ton tai -> co the gop (nhung tam thoi tao moi theo form nay)
                
                db.add_hop_dong({
                    "ma_hop_dong": ma_hd, "khach_hang": kh, "cif": cif_text, "don_vi_thu_huong": dvth,
                    "so_hd": so_hd, "ngay_hop_dong": ngay_hd, "gia_tri_hd": gtri, "ngay_ket_thuc_hd": ngay_kt,
                    "loai_tu": "mot_lan" if is_mot_lan else "khau_tru_dot",
                    "loai_gia_tri_kt": loai_kt,
                    "pct_khau_tru": pct, "phong_phu_trach": phong, "ghi_chu": ghi_chu, "khe_uoc_vay": khe_uoc_vay
                }, user)
                
                ma_gn = db.generate_ma_giai_ngan(kh, year=ngay_gn.year)
                db.add_tam_ung({
                    "ma_giai_ngan": ma_gn, "ma_hop_dong": ma_hd,
                    "so_tien_tu": tu_amt, "ngay_giai_ngan": ngay_gn, "ghi_chu": "Giải ngân lần 1"
                }, user)
                
                st.success(f"✅ Tạo thành công Hợp đồng: {ma_hd} và Đợt giải ngân: {ma_gn}")


# ============ PAGE: DUYET HD CHO ============
def page_duyet_hd():
    st.title("⏳ Duyệt Hóa đơn chờ")
    pending = db.list_hoa_don(status="pending")
    if not pending:
        st.success("Không có HĐ chờ duyệt.")
        return
    for h in pending:
        with st.container(border=True):
            c1, c2 = st.columns([3, 1])
            with c1:
                st.write(f"**HD {h['so_hd']}** | Ngày: {fmt_date(h['ngay_hd'])} | MST: `{h['mst_ban'] or '-'}`")
                st.write(f"**Người bán:** {h['ten_ban'] or '-'}")
                st.write(f"**Số tiền:** {fmt_vnd(h['tien_truoc_vat'])} | **VAT:** {fmt_vnd(h['vat'])} | **Tổng:** {fmt_vnd(h['tong_cong'])}")
                if h['file_src'] and os.path.exists(h['file_src']):
                    with open(h['file_src'], "rb") as f:
                        st.download_button("📥 Tải file đính kèm", f, file_name=os.path.basename(h['file_src']), key=f"dl_{h['id']}")
            with c2:
                if st.button("✓ Duyệt", key=f"app_{h['id']}", type="primary", use_container_width=True):
                    db.approve_hoa_don(h["id"], user)
                    st.rerun()
                if st.button("✗ Từ chối", key=f"rej_{h['id']}", use_container_width=True):
                    db.reject_hoa_don(h["id"], user, "Từ chối")
                    st.rerun()


# ============ PAGE: IMPORT EXCEL ============
def page_import_excel():
    st.title("📥 Import Danh Sách Hợp Đồng Từ Excel")
    st.write("Chức năng này giúp tạo hàng loạt Hợp đồng & Đợt Tạm ứng đầu tiên từ file Excel.")
    
    # Tạo File Mẫu
    import io
    import pandas as pd
    
    cols = [
        "Khách hàng (Bắt buộc)",
        "Mã CIF",
        "Đơn vị thụ hưởng",
        "Số Hợp đồng",
        "Ngày Hợp đồng (DD/MM/YYYY)",
        "Giá trị HĐ (VND)",
        "Ngày kết thúc HĐ (DD/MM/YYYY) (Bắt buộc)",
        "Tỷ lệ Khấu trừ (%) (Bắt buộc)",
        "Phòng phụ trách",
        "Khế ước vay",
        "Số tiền Giải ngân đợt 1 (VND) (Bắt buộc)",
        "Ngày giải ngân (DD/MM/YYYY) (Bắt buộc)",
        "Ghi chú"
    ]
    df_template = pd.DataFrame(columns=cols)
    
    towrite = io.BytesIO()
    df_template.to_excel(towrite, index=False, engine='openpyxl')
    towrite.seek(0)
    
    st.download_button(
        label="⬇️ Tải File Excel Mẫu",
        data=towrite,
        file_name="Template_Import_HopDong.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.divider()
    st.subheader("Tải lên File Excel đã điền")
    
    uploaded_file = st.file_uploader("Chọn file Excel (.xlsx)", type=["xlsx"])
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file, dtype=str)
            df = df.fillna("")
            
            # Đổi tên cột cho dễ xử lý nội bộ nếu cần, hoặc dùng index/tên cột đầy đủ
            st.write(f"**Đã tải lên {len(df)} dòng dữ liệu.** Xem trước:")
            st.dataframe(df.head(5))
            
            if st.button("🚀 Bắt đầu Import", type="primary"):
                user = st.session_state["user"]
                username = user["username"]
                
                success = 0
                failed = 0
                errors = []
                
                progress_bar = st.progress(0)
                
                for idx, row in df.iterrows():
                    try:
                        kh = str(row.get("Khách hàng (Bắt buộc)", "")).strip()
                        if not kh:
                            failed += 1
                            errors.append(f"Dòng {idx+2}: Thiếu tên Khách hàng.")
                            continue
                            
                        tien_tu_str = str(row.get("Số tiền Giải ngân đợt 1 (VND) (Bắt buộc)", "")).strip()
                        tien_tu = parse_vnd_input(tien_tu_str)
                        if not tien_tu:
                            failed += 1
                            errors.append(f"Dòng {idx+2}: Thiếu/sai định dạng Số tiền giải ngân.")
                            continue
                            
                        tlkt_str = str(row.get("Tỷ lệ Khấu trừ (%) (Bắt buộc)", "")).strip()
                        if not tlkt_str:
                            failed += 1
                            errors.append(f"Dòng {idx+2}: Thiếu Tỷ lệ khấu trừ.")
                            continue
                        try:
                            tlkt = float(tlkt_str.replace("%","").replace(",","."))
                        except:
                            failed += 1
                            errors.append(f"Dòng {idx+2}: Sai định dạng Tỷ lệ khấu trừ.")
                            continue
                            
                        ngay_kt_str = str(row.get("Ngày kết thúc HĐ (DD/MM/YYYY) (Bắt buộc)", "")).strip()
                        ngay_kt = None
                        try:
                            if ngay_kt_str: ngay_kt = datetime.strptime(ngay_kt_str, "%d/%m/%Y").date()
                        except: pass
                        if not ngay_kt:
                            failed += 1
                            errors.append(f"Dòng {idx+2}: Thiếu/sai định dạng Ngày kết thúc HĐ.")
                            continue
                            
                        ngay_gn_str = str(row.get("Ngày giải ngân (DD/MM/YYYY) (Bắt buộc)", "")).strip()
                        ngay_gn = None
                        try:
                            if ngay_gn_str: ngay_gn = datetime.strptime(ngay_gn_str, "%d/%m/%Y").date()
                        except: pass
                        if not ngay_gn:
                            failed += 1
                            errors.append(f"Dòng {idx+2}: Thiếu/sai định dạng Ngày giải ngân.")
                            continue
                            
                        ngay_hd_str = str(row.get("Ngày Hợp đồng (DD/MM/YYYY)", "")).strip()
                        ngay_hd = None
                        try:
                            if ngay_hd_str: ngay_hd = datetime.strptime(ngay_hd_str, "%d/%m/%Y").date()
                        except: pass
                            
                        # Tao ma HD
                        ma_hd = db.generate_ma_hop_dong(kh)
                        db.add_hop_dong(
                            ma_hd=ma_hd,
                            khach_hang=kh,
                            cif=str(row.get("Mã CIF", "")).strip() or None,
                            don_vi_thu_huong=str(row.get("Đơn vị thụ hưởng", "")).strip() or None,
                            so_hd=str(row.get("Số Hợp đồng", "")).strip() or None,
                            gia_tri_hd=parse_vnd_input(row.get("Giá trị HĐ (VND)", "")) or None,
                            ngay_hd=ngay_hd,
                            ngay_ket_thuc_hd=ngay_kt,
                            loai_tu="khau_tru_dot", # default
                            loai_gia_tri_kt=None,
                            pct_khau_tru=tlkt,
                            phong_phu_trach=str(row.get("Phòng phụ trách", "")).strip() or "Khác",
                            ghi_chu=str(row.get("Ghi chú", "")).strip() or None,
                            khe_uoc_vay=str(row.get("Khế ước vay", "")).strip() or None,
                            created_by=username
                        )
                        
                        # Tao ma GN & add Tam Ung
                        ma_gn = db.generate_ma_giai_ngan(ma_hd)
                        db.add_tam_ung(
                            ma_giai_ngan=ma_gn,
                            ma_hop_dong=ma_hd,
                            so_tien_tu=tien_tu,
                            ngay_giai_ngan=ngay_gn,
                            ghi_chu=f"Import từ Excel - {str(row.get('Ghi chú', '')).strip()}",
                            created_by=username
                        )
                        
                        success += 1
                    except Exception as e:
                        failed += 1
                        errors.append(f"Dòng {idx+2}: Lỗi hệ thống ({str(e)})")
                    
                    progress_bar.progress(int(((idx+1) / len(df)) * 100))
                
                # Hoan tat
                progress_bar.empty()
                if success > 0:
                    st.success(f"🎉 Đã import thành công {success} Hợp đồng!")
                    db.log_action("IMPORT_EXCEL", "hop_dong", f"{success} HD", username, f"Import thanh cong {success} HD tu file")
                if failed > 0:
                    st.error(f"❌ Thất bại {failed} dòng. Chi tiết:")
                    for err in errors:
                        st.write(f"- {err}")
                        
        except Exception as e:
            st.error(f"Lỗi đọc file Excel: {e}")

# ============ PAGE: UPLOAD THEO LO ============
def page_upload_lo():
    st.title("📥 Up hóa đơn điện tử theo lô")
    st.write("Tải lên hàng loạt file XML hóa đơn và xuất ra mẫu Excel hệ thống.")
    
    phong_filter = None if is_qttd else uinfo["dept"]
    
    mode = st.radio("Chế độ chọn", ["Theo Hợp đồng", "Theo Ngày bổ sung"], horizontal=True)
    
    selected_contract = None
    date_choice = None
    
    if mode == "Theo Hợp đồng":
        all_rows = db.list_hop_dong(phong=phong_filter)
        if not all_rows:
            st.warning("Chưa có Hợp đồng nào trong hệ thống.")
            return
            
        options = {}
        for r in all_rows:
            s, tt = compute_hd_status(r)
            kh_short = shorten_company_name(r.get('khach_hang', ''))
            dv_short = shorten_company_name(r.get('don_vi_thu_huong', ''))
            so_hd = r.get('so_hd') or "Chưa rõ"
            tien = f"{fmt_vnd_dot(s['tong_giai_ngan'])}đ" if s['tong_giai_ngan'] else "0đ"
            
            display_name = f"{kh_short} - HĐ: {so_hd} - GN: {tien}"
            if dv_short: display_name += f" - ĐVTH: {dv_short}"
            options[display_name] = r
            
        choice = st.selectbox("📌 Chọn Hợp đồng để xem Hóa đơn:", list(options.keys()))
        selected_contract = options[choice]
        hoadons = db.list_hoa_don(ma_hop_dong=selected_contract["ma_hop_dong"])
        if not hoadons:
            st.info("Chưa có hóa đơn nào được tải lên cho Hợp đồng này. Vui lòng bổ sung hóa đơn ở menu Danh sách Hợp đồng trước.")
            return
    else:
        dates = db.get_distinct_upload_dates(phong=phong_filter)
        if not dates:
            st.warning("Chưa có Hóa đơn nào được upload.")
            return
        
        date_choice = st.selectbox("📌 Chọn Ngày bổ sung:", dates, format_func=lambda d: fmt_date(d).replace("/", "-"))
        hoadons = db.get_hoadons_by_date_str(date_choice, phong=phong_filter)
        if not hoadons:
            st.info("Không có hóa đơn nào trong ngày này.")
            return
        
    st.write("### Chọn các hóa đơn để xuất Excel")
    
    selected_hds = []
    
    toggle_key = f"chk_select_all_{mode}_{selected_contract['ma_hop_dong'] if selected_contract else date_choice}"
    def toggle_select_all():
        val = st.session_state.get(toggle_key, False)
        for h in hoadons:
            st.session_state[f"chk_export_{h['id']}"] = val
            
    st.checkbox("Chọn tất cả", key=toggle_key, on_change=toggle_select_all)
    
    for h in hoadons:
        col1, col2, col3 = st.columns([1, 7, 2], vertical_alignment="center")
        is_checked = col1.checkbox("Chọn", key=f"chk_export_{h['id']}", label_visibility="collapsed")
        
        uploaded = h['uploaded_at'] if 'uploaded_at' in h.keys() else None
        ngay_bs = fmt_date(str(uploaded).split(' ')[0]) if uploaded else '-'
        if mode == "Theo Hợp đồng":
            col2.markdown(f"**Ngày BS:** <span style='color:blue'>{ngay_bs}</span> | **Số HĐ:** {h['so_hd']} | **Ngày HĐ:** {fmt_date(h['ngay_hd'])} | **ĐVTH:** {h['ten_ban']}", unsafe_allow_html=True)
        else:
            contract_info = db.get_hop_dong(h['ma_hop_dong'])
            kh_short = shorten_company_name(contract_info['khach_hang']) if contract_info else "Khách vãng lai"
            col2.markdown(f"**HĐ:** <span style='color:green'>{kh_short}</span> | **Số HĐ:** {h['so_hd']} | **Ngày HĐ:** {fmt_date(h['ngay_hd'])} | **ĐVTH:** {h['ten_ban']}", unsafe_allow_html=True)
            
        col3.write(f"**Tổng tiền:** {fmt_vnd(h['tong_cong'])}")
        if is_checked:
            selected_hds.append(h)
            
    if len(selected_hds) > 0 and st.button("📥 Tạo file Excel", type="primary"):
        with st.spinner("Đang xử lý..."):
            parsed_data = []
            for h in selected_hds:
                d_extra = {}
                if h['file_src'] and os.path.exists(h['file_src']):
                    with open(h['file_src'], "rb") as f:
                        d_extra = parsers.parse_file(f)
                
                contract = db.get_hop_dong(h['ma_hop_dong'])
                if not contract: continue
                
                pct = 1.0 if contract["loai_tu"] == "mot_lan" else (contract["pct_khau_tru"] or 0) / 100.0
                loai_kt = "Trước VAT" if contract["loai_tu"] == "mot_lan" else (contract["loai_gia_tri_kt"] or "Trước VAT")
                gia_tri_tinh = h['tong_cong'] if loai_kt == "Sau VAT" else h['tien_truoc_vat']
                can_tru = (gia_tri_tinh or 0) * pct
                
                tus = db.list_tam_ung(contract["ma_hop_dong"])
                ngay_gn_dau = tus[0]["ngay_giai_ngan"] if tus else date.today()
                
                parsed_data.append({
                    "so_hd": h["so_hd"],
                    "ngay_hd": h["ngay_hd"],
                    "mst_ban": h["mst_ban"],
                    "ten_ban": h["ten_ban"],
                    "tong_cong": h["tong_cong"],
                    "mau_so_hd": d_extra.get("mau_so_hd"),
                    "ky_hieu_hd": d_extra.get("ky_hieu_hd"),
                    "can_tru": can_tru,
                    "cif": contract['cif'] if 'cif' in contract.keys() and contract['cif'] else "",
                    "ngay_gn_dau": ngay_gn_dau,
                    "khe_uoc_vay": contract['khe_uoc_vay'] if 'khe_uoc_vay' in contract.keys() and contract['khe_uoc_vay'] else ""
                })
                
            try:
                # Tim file Template.xlsx
                t1_paths = [
                    "Template.xlsx",
                    os.path.join(os.path.dirname(__file__), "Template.xlsx"),
                    os.path.join(os.path.dirname(__file__), "..", "Template.xlsx"),
                    os.path.join(os.getcwd(), "outputs", "Template.xlsx"),
                    os.path.join(os.getcwd(), "Template.xlsx")
                ]
                template_path = next((p for p in t1_paths if os.path.exists(p)), "Template.xlsx")
                
                wb = openpyxl.load_workbook(template_path)
                ws = wb.active
                
                val_f = ws.cell(row=8, column=6).value
                
                start_row = 8
                for i, d in enumerate(parsed_data):
                    r = start_row + i
                    # A: CIF
                    ws.cell(row=r, column=1, value=d['cif'])
                    # B: Mẫu số hóa đơn
                    ws.cell(row=r, column=2, value=d.get("mau_so_hd") or "1")
                    # C: Ký hiệu
                    ws.cell(row=r, column=3, value=d.get("ky_hieu_hd") or "")
                    # D: Loại Hóa đơn
                    ws.cell(row=r, column=4, value="01HDDT")
                    # E: Số hóa đơn
                    ws.cell(row=r, column=5, value=d.get("so_hd") or "")
                    # F: Số HĐ xác thực (Cột F)
                    ws.cell(row=r, column=6, value=val_f)
                    # G: Giá trị hóa đơn
                    ws.cell(row=r, column=7, value=d.get("tong_cong") or 0)
                    # H: Loại tiền
                    ws.cell(row=r, column=8, value="VND")
                    # I: Tỷ giá
                    ws.cell(row=r, column=9, value=1)
                    # J: Ngày hóa đơn
                    ngay = d.get("ngay_hd")
                    if isinstance(ngay, str):
                        try: ngay = datetime.strptime(ngay, "%Y-%m-%d").date()
                        except: pass
                    ws.cell(row=r, column=10, value=ngay.strftime("%d/%m/%Y") if isinstance(ngay, date) else str(ngay))
                    # K: MST
                    ws.cell(row=r, column=11, value=d.get("mst_ban") or "")
                    # L: Đơn vị phát hành
                    ws.cell(row=r, column=12, value=d.get("ten_ban") or "")
                    # M: Ngân hàng
                    ws.cell(row=r, column=13, value="135")
                    # N: Ngày tháng năm giải ngân
                    ws.cell(row=r, column=14, value=fmt_date(d['ngay_gn_dau']))
                    # O: Số tiền giải ngân (quy đổi VNĐ) - Giá trị thu hồi tạm ứng
                    ws.cell(row=r, column=15, value=d.get("can_tru") or 0)
                    # P: Số tài khoản tiền vay
                    ws.cell(row=r, column=16, value=d['khe_uoc_vay'])
                    # Q: Loại chứng từ
                    ws.cell(row=r, column=17, value="01HDDT")
                
                    for c in range(1, 18):
                        source_cell = ws.cell(row=8, column=c)
                        target_cell = ws.cell(row=r, column=c)
                        if source_cell.has_style:
                            target_cell.font = copy.copy(source_cell.font)
                            target_cell.border = copy.copy(source_cell.border)
                            target_cell.fill = copy.copy(source_cell.fill)
                            target_cell.number_format = source_cell.number_format
                            target_cell.protection = copy.copy(source_cell.protection)
                            target_cell.alignment = copy.copy(source_cell.alignment)
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # --- PROCESS SECOND EXCEL: BANG KE HOA DON ---
                t2_paths = [
                    "template_hoa_don.xlsx",
                    os.path.join(os.path.dirname(__file__), "template_hoa_don.xlsx"),
                    os.path.join(os.path.dirname(__file__), "..", "template_hoa_don.xlsx"),
                    os.path.join(os.getcwd(), "outputs", "template_hoa_don.xlsx"),
                    os.path.join(os.getcwd(), "template_hoa_don.xlsx")
                ]
                template_hd_path = next((p for p in t2_paths if os.path.exists(p)), None)
                
                if template_hd_path:
                    wb2 = openpyxl.load_workbook(template_hd_path)
                    ws2 = wb2.active
                    
                    start_row2 = 2
                    for i, d in enumerate(parsed_data):
                        r = start_row2 + i
                        # 1: STT
                        ws2.cell(row=r, column=1, value=i+1)
                        # 2: Số HĐ
                        ws2.cell(row=r, column=2, value=d.get("so_hd") or "")
                        # 3: Ký hiệu
                        ws2.cell(row=r, column=3, value=d.get("ky_hieu_hd") or "")
                        # 4: Ngày HĐ
                        ngay = d.get("ngay_hd")
                        if isinstance(ngay, str):
                            try: ngay = datetime.strptime(ngay, "%Y-%m-%d").date()
                            except: pass
                        ws2.cell(row=r, column=4, value=ngay.strftime("%d/%m/%Y") if isinstance(ngay, date) else str(ngay))
                        # 5: Số tiền HĐ (VNĐ)
                        ws2.cell(row=r, column=5, value=d.get("tong_cong") or 0)
                        # 6: Tên đơn vị phát hành
                        ws2.cell(row=r, column=6, value=d.get("ten_ban") or "")
                        # 7: MST
                        ws2.cell(row=r, column=7, value=d.get("mst_ban") or "")
                        
                        if r > start_row2:
                            for c in range(1, 8):
                                source_cell = ws2.cell(row=start_row2, column=c)
                                target_cell = ws2.cell(row=r, column=c)
                                if source_cell.has_style:
                                    target_cell.font = copy.copy(source_cell.font)
                                    target_cell.border = copy.copy(source_cell.border)
                                    target_cell.fill = copy.copy(source_cell.fill)
                                    target_cell.number_format = source_cell.number_format
                                    target_cell.protection = copy.copy(source_cell.protection)
                                    target_cell.alignment = copy.copy(source_cell.alignment)
                    
                    output2 = io.BytesIO()
                    wb2.save(output2)
                    output2.seek(0)
                else:
                    output2 = None
                
                st.success(f"✅ Đã xử lý {len(parsed_data)} hóa đơn thành công!")
                
                suffix = selected_contract['ma_hop_dong'] if selected_contract else str(date_choice)
                
                col_btn1, col_btn2 = st.columns(2)
                col_btn1.download_button(
                    label="📥 Tải file Upload Lô",
                    data=output,
                    file_name=f"UploadLo_{suffix}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                if output2:
                    col_btn2.download_button(
                        label="📥 Tải Bảng kê Hóa đơn",
                        data=output2,
                        file_name=f"BangKeHD_{suffix}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    col_btn2.warning("Không tìm thấy mẫu Bảng kê hóa đơn.")
            except Exception as e:
                st.error(f"Lỗi khi xử lý Template Excel: {e}")

# ============ ROUTE ============
if page.startswith("📊 Dashboard"):
    page_dashboard()
elif page == "➕ Thêm Hợp đồng Tạm ứng":
    page_them_hop_dong()
elif page == "📥 Import Hợp đồng (Excel)":
    page_import_excel()
elif page in ("💼 Quản lý Hợp đồng", "💼 Danh sách Hợp đồng"):
    page_hop_dong()
elif page == "⏳ Duyệt Hóa đơn" or page == "✅ Duyệt HĐ chờ":
    page_duyet_hd()
elif page == "📥 Up hóa đơn theo lô":
    page_upload_lo()
elif page == "📦 Export & Drive":
    st.title("📦 Xuất báo cáo Excel")
    st.write("Xuất toàn bộ dữ liệu Hợp đồng, Giải ngân, Hóa đơn ra file Excel theo chuẩn.")
    if st.button("📥 Xuất file Excel", type="primary"):
        with st.spinner("Đang xử lý dữ liệu..."):
            try:
                path = exporter.export_excel()
                with open(path, "rb") as f:
                    st.download_button(
                        label="Tải file Excel về máy",
                        data=f,
                        file_name=os.path.basename(path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                st.success("Tạo file thành công!")
            except Exception as e:
                st.error(f"Lỗi khi xuất file: {e}")
elif page == "🗂 Audit log":
    st.dataframe(pd.DataFrame([dict(r) for r in db.recent_audit(100)]), use_container_width=True)
